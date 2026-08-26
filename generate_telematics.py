import csv
import json
import random
from datetime import datetime, timedelta, timezone
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NUM_RECORDS = 100
VEHICLE_IDS = [f"VEH-CT-{100 + i}" for i in range(5)]
START_TIME = datetime.now(timezone.utc) - timedelta(days=1)

dataset = []
for i in range(NUM_RECORDS):
    record_time = START_TIME + timedelta(minutes=i * 5)
    is_harsh = random.choices([True, False], weights=[0.08, 0.92])[0]
    battery = round(random.uniform(15.0, 100.0), 1)
    speed = round(random.uniform(0.0, 110.0), 1)

    if battery < 20:
        bat_status = "CRITICAL"
    elif battery < 40:
        bat_status = "LOW"
    else:
        bat_status = "NORMAL"

    dataset.append({
        "telemetry_id": f"TEL-{10000 + i}",
        "vehicle_id": random.choice(VEHICLE_IDS),
        "timestamp": record_time.strftime("%Y-%m-%d %H:%M:%S"),
        "latitude": round(random.uniform(-33.8850, -33.8500), 6),
        "longitude": round(random.uniform(151.1900, 151.2200), 6),
        "speed_kmh": speed,
        "battery_level_pct": battery,
        "battery_status": bat_status,
        "harsh_braking_event": "YES" if is_harsh else "NO",
    })

# Save JSON
with open("ctrack_telematics_synthetic.json", "w") as jf:
    json.dump(dataset, jf, indent=4)

# Save CSV
with open("ctrack_telematics_synthetic.csv", "w", newline="") as cf:
    writer = csv.DictWriter(cf, fieldnames=dataset[0].keys())
    writer.writeheader()
    writer.writerows(dataset)

# Save Formatted Excel
wb = openpyxl.Workbook()
ws_summary = wb.active
ws_summary.title = "Executive Summary"
ws_summary.views.sheetView[0].showGridLines = True

ws_summary.merge_cells("A1:E2")
title_cell = ws_summary["A1"]
title_cell.value = "Ctrack Telematics Synthetic Dataset Summary (AI-3)"
title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="1B365D")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws_summary["A4"] = "Prepared By:"
ws_summary["B4"] = "Sushil SAPKOTA (Assignee)"
ws_summary["A5"] = "Project / Task:"
ws_summary["B5"] = "Ctrack AI Telematics Assistant / Task AI-3"
ws_summary["A6"] = "Sprint:"
ws_summary["B6"] = "AI Sprint 1 (25 Aug – 22 Sep)"

for row in range(4, 7):
    ws_summary[f"A{row}"].font = Font(bold=True, color="1B365D")

ws_summary.merge_cells("A9:D9")
kpi_title = ws_summary["A9"]
kpi_title.value = "Dataset Key Metrics Summary"
kpi_title.font = Font(bold=True, color="FFFFFF", size=12)
kpi_title.fill = PatternFill("solid", fgColor="2C3E50")
kpi_title.alignment = Alignment(horizontal="left", vertical="center")

headers_kpi = ["Metric Description", "Value", "Unit / Format", "Target / Note"]
for col_num, h in enumerate(headers_kpi, 1):
    c = ws_summary.cell(row=10, column=col_num, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4A6572")
    c.alignment = Alignment(horizontal="center", vertical="center")

kpi_data = [
    (
        "Total Telemetry Logs",
        "=COUNTA('Telematics Data'!A2:A101)",
        "Records",
        "Target: 100 mock events",
    ),
    (
        "Average Fleet Speed",
        "=AVERAGE('Telematics Data'!F2:F101)",
        "km/h",
        "City & Highway mix",
    ),
    (
        "Average Battery Level",
        "=AVERAGE('Telematics Data'!G2:G101)",
        "%",
        "Fleet battery health",
    ),
    (
        "Harsh Braking Incidents",
        '=COUNTIF(\'Telematics Data\'!I2:I101, "YES")',
        "Incidents",
        "Safety anomaly flag",
    ),
]

thin_border = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)

for row_idx, data_row in enumerate(kpi_data, 11):
    for col_idx, val in enumerate(data_row, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border

ws_summary.column_dimensions["A"].width = 30
ws_summary.column_dimensions["B"].width = 35
ws_summary.column_dimensions["C"].width = 18
ws_summary.column_dimensions["D"].width = 30

ws_data = wb.create_sheet(title="Telematics Data")
ws_data.views.sheetView[0].showGridLines = True

headers_data = [
    "Telemetry ID",
    "Vehicle ID",
    "Timestamp (UTC)",
    "Latitude",
    "Longitude",
    "Speed (km/h)",
    "Battery Level (%)",
    "Battery Status",
    "Harsh Braking",
]

for col_num, header in enumerate(headers_data, 1):
    cell = ws_data.cell(row=1, column=col_num, value=header)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1B365D")
    cell.alignment = Alignment(horizontal="center", vertical="center")

fill_harsh = PatternFill("solid", fgColor="FFC7CE")
font_harsh = Font(color="9C0006", bold=True)
fill_low = PatternFill("solid", fgColor="FFEB9C")
font_low = Font(color="9C6500", bold=True)

for row_idx, rec in enumerate(dataset, 2):
    ws_data.cell(row=row_idx, column=1, value=rec["telemetry_id"]).alignment = (
        Alignment(horizontal="center")
    )
    ws_data.cell(row=row_idx, column=2, value=rec["vehicle_id"]).alignment = (
        Alignment(horizontal="center")
    )
    ws_data.cell(row=row_idx, column=3, value=rec["timestamp"]).alignment = (
        Alignment(horizontal="center")
    )
    ws_data.cell(row=row_idx, column=4, value=rec["latitude"]).alignment = (
        Alignment(horizontal="right")
    )
    ws_data.cell(row=row_idx, column=5, value=rec["longitude"]).alignment = (
        Alignment(horizontal="right")
    )
    ws_data.cell(row=row_idx, column=6, value=rec["speed_kmh"]).alignment = (
        Alignment(horizontal="right")
    )
    ws_data.cell(row=row_idx, column=7, value=rec["battery_level_pct"]).alignment = (
        Alignment(horizontal="right")
    )

    c_stat = ws_data.cell(row=row_idx, column=8, value=rec["battery_status"])
    c_stat.alignment = Alignment(horizontal="center")
    if rec["battery_status"] in ["LOW", "CRITICAL"]:
        c_stat.fill = fill_low
        c_stat.font = font_low

    c_hb = ws_data.cell(row=row_idx, column=9, value=rec["harsh_braking_event"])
    c_hb.alignment = Alignment(horizontal="center")
    if rec["harsh_braking_event"] == "YES":
        c_hb.fill = fill_harsh
        c_hb.font = font_harsh

    for col_idx in range(1, 10):
        ws_data.cell(row=row_idx, column=col_idx).border = thin_border

for col in ws_data.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_data.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save("ctrack_telematics_synthetic.xlsx")
print("Generated JSON, CSV, and XLSX successfully.")